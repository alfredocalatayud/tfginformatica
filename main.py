import time
import pandas as pd
import requests
from progress.bar import Bar
import openpyxl as op

# Diccionario de descripciones de estados HTTP
descripciones_http = {
    100: "Continuar",
    101: "Cambiando protocolos",
    102: "Procesando",
    200: "OK",
    201: "Creado",
    202: "Aceptado",
    203: "Información no autoritativa",
    204: "Sin contenido",
    205: "Restablecer contenido",
    206: "Contenido parcial",
    207: "Estado multiestado",
    208: "Estado ya informado",
    226: "IM utilizado",
    300: "Múltiples opciones",
    301: "Movido permanentemente",
    302: "Encontrado",
    303: "Vea otros",
    304: "No modificado",
    305: "Usar proxy",
    307: "Redirección temporal",
    308: "Redirección permanente",
    400: "Solicitud incorrecta",
    401: "No autorizado",
    402: "Pago requerido",
    403: "Prohibido",
    404: "No encontrado",
    405: "Método no permitido",
    406: "No aceptable",
    407: "Autenticación de proxy requerida",
    408: "Solicitud de tiempo de espera",
    409: "Conflicto",
    410: "Ya no disponible",
    411: "Longitud requerida",
    412: "Falló la precondición",
    413: "Solicitud de entidad demasiado grande",
    414: "URI demasiado larga",
    415: "Tipo de medio no soportado",
    416: "Rango solicitado no satisfactorio",
    417: "Falló la expectativa",
    418: "Soy una tetera",
    421: "Solicitud mal dirigida",
    422: "Entidad no procesable",
    423: "Bloqueado",
    424: "Fallo dependiente",
    426: "Actualización requerida",
    428: "Precondición obligatoria",
    429: "Demasiadas solicitudes",
    431: "Campos de encabezado de solicitud demasiado grandes",
    451: "No disponible por razones legales",
    500: "Error interno del servidor",
    501: "No implementado",
    502: "Puerta de enlace incorrecta",
    503: "Servicio no disponible",
    504: "Tiempo de espera de la puerta de enlace",
    505: "Versión HTTP no compatible",
    506: "Variante también negocia",
    507: "Espacio insuficiente de almacenamiento",
    508: "Ciclo detectado",
    510: "No extendido",
    511: "Autenticación de red requerida",
}

n_urls = 0
n_urls_off = 0
n_intentos = 0
n_retries = 3  # Número máximo de reintentos
retry_delay = 5  # Tiempo de espera entre reintentos en segundos

urls_off = []
l_http_status = []

db_xlsx = "ayuntamientos_web_v2.xlsx"

try:
    df = pd.read_excel(db_xlsx)
except FileNotFoundError:
    print("El archivo {} no se encontró.".format(db_xlsx))
    exit()
except Exception as e:
    print("Ocurrió un error al leer el archivo: {}".format(e))
    exit()

urls = df["WebAyuntamiento"]
ayuntamientos = df["NombreAyuntamiento"]

n_urls = len(urls)

bar = Bar('Probando URLs:', max=n_urls)

for url in urls:
    bar.next()
    for _ in range(n_retries):
        try:
            response = requests.get(url, timeout=10)  # Ajusta el tiempo de espera según sea necesario
            http_status = response.status_code
            l_http_status.append(http_status)

            if http_status != 200:
                n_urls_off += 1
                description = descripciones_http.get(http_status, "Estado desconocido")
                urls_off.append((url, http_status, description))
            break  # Sale del bucle de reintento si la solicitud es exitosa
        except requests.exceptions.RequestException as e:
            n_intentos += 1
            time.sleep(retry_delay)  # Espera antes de reintentar

    n_urls += 1

    #if n_urls >= 100:
     #   break

bar.finish()

with open('informe.txt', 'w') as txtfile:
    txtfile.write("Informe de Estados HTTP:\n\n")
    txtfile.write("Número de URLs revisadas: {}\n".format(n_urls))
    txtfile.write("Número de URLs con estado distinto de 200: {}\n".format(n_urls_off))
    txtfile.write("Número de reintentos de acceso: {}\n\n".format(n_intentos))

    if n_urls_off > 0:
        txtfile.write("URLs con estado distinto de 200:\n")
        for url, status, description in urls_off:
            ayuntamiento = df.loc[df["WebAyuntamiento"] == url, "NombreAyuntamiento"].values[0]
            txtfile.write("Ayuntamiento: {}\nURL: {}\nEstado HTTP: {} ({})\n\n".format(ayuntamiento, url, status, description))

print("Informe generado exitosamente en 'informe.txt'")

# Abre el archivo Excel
try:
    workbook = op.load_workbook(db_xlsx)
except FileNotFoundError:
    print(f"El archivo {db_xlsx} no se encontró.")
    exit()
except Exception as e:
    print(f"Ocurrió un error al leer el archivo: {str(e)}")
    exit()

# Obtiene la hoja activa del archivo
sheet = workbook.active

# Especifica el nombre de la columna en el DataFrame
column_name = "HTTPStatus"

# Verifica si la columna ya existe, de lo contrario, créala
if column_name not in sheet[1]:
    sheet.insert_cols(idx=6, amount=1)  # Inserta la columna en la posición deseada
    sheet.cell(row=1, column=6, value=column_name)  # Establece el encabezado de la columna

# Escribe los valores de l_http_status en la columna especificada
for idx, http_status in enumerate(l_http_status, start=2):
    cell = sheet.cell(row=idx, column=6)
    cell.value = http_status

# Guarda el archivo Excel actualizado
workbook.save(db_xlsx)
workbook.close()

print("Valores de estado agregados exitosamente en la columna '{}' de '{}'.".format(column_name, db_xlsx))
